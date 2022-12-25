import shutil
import sqlite3
import os
from datetime import date, datetime

import openpyxl


DATE = date.today() # сегодняшняя дата (объект)
CONNECT = r'\\cronosx1\New folder\УВБ\Отдел корпоративной защиты\candidates.db' #     файл базы данных
WORK_DIR = r'\\cronosx1\New folder\УВБ\Отдел корпоративной защиты\Кандидаты\\'    # рабочая папка кандидатов
DESTINATION = r'\\cronosx1\New folder\УВБ\Отдел корпоративной защиты\Персонал\Персонал-2\\'   # архивная папка
INFO_FILE = r'\\cronosx1\New folder\УВБ\Отдел корпоративной защиты\Кандидаты\Запросы по работникам.xlsx'  # запросы
MAIN_FILE = r'\\cronosx1\New folder\УВБ\Отдел корпоративной защиты\Кандидаты\Кандидаты.xlsm'  # файл кандидатов
# поля в таблице candidates базы данных candidates:
SQL_CAND = 'staff, department, full_name, last_name, birthday, birth_place, country, series_passport, ' \
           'number_passport, date_given, snils, inn, reg_address, live_address, phone, email, education, ' \
           'check_work_place, check_passport, check_debt, check_bankruptcy, check_bki, check_affiliation, ' \
           'check_internet, check_cronos, check_cross, resume, date_check, officer'
# поля в таблице registry базы данных candidates:
SQL_REG = 'fio, birthday, staff, checks, recruiter, date_in, officer, date_out, result, final_date, url'
# поля в таблице inquiry базы данных candidates:
SQL_INQ = 'full_name, birthday, staff, period, info, firm, date_inq'
# тестовые:
# CONNECT = r'C:\Users\ubuntu\Documents\Отдел корпоративной защиты\candidates.db'  #     файл базы данных
# WORK_DIR = r'C:\Users\ubuntu\Documents\Отдел корпоративной защиты\Кандидаты\\'  # рабочая папка
# DESTINATION = r'C:\Users\ubuntu\Documents\Отдел корпоративной защиты\Персонал\Персонал-2\\' # архивная папка
# MAIN_FILE = r'C:\Users\ubuntu\Documents\Отдел корпоративной защиты\Кандидаты\Кандидаты.xlsm'   # файл кандидатов
# INFO_FILE = r'C:\Users\ubuntu\Documents\Отдел корпоративной защиты\Кандидаты\Запросы по работникам.xlsx' # запросы


def range_row(sheet) -> list: # получаем список номеров строк соответствующих сегодняшней дате
    row_num = []
    for cell in sheet:
        for c in cell:
            if isinstance (c.value, datetime):
                if c.value.strftime('%Y-%m-%d') == DATE.strftime('%Y-%m-%d'):
                    row_num.append(c.row)
            elif str(c.value).strip() == DATE.strftime('%d.%m.%Y'):
                row_num.append(c.row)
    return row_num

def dir_range(sheet, row_num) -> list:  # получаем список папок в рабочей директории
    fio = [sheet['B' + str(i)].value.strip().lower() for i in row_num]
    subdir = [sub for sub in os.listdir(WORK_DIR) if sub.lower().strip() in fio]
    return subdir

def file_range(subdir) -> list: # получаем список путей к файлам Заключений
    name_path = list(filter(None, []))
    for f in subdir:
        subdir_path = f"{WORK_DIR[0:-1] + f}\\"
        for file_name in os.listdir(subdir_path):
            if file_name.startswith("Заключение") and file_name.endswith("xlsm"):
                name_path.append(os.path.join(WORK_DIR, subdir_path, file_name))
    return name_path

def check_types(check) -> list: # преобразование типов данных
    view = []
    for c in check:
        if isinstance(c, datetime):  # преобразование datetime в строку
            view.append(c.strftime('%Y-%m-%d'))
        else:
            view.append(str(c).strip())    # преобразование в строку
    return view

def parse_excel(path_files) -> list: # получаем список кортежей с данными заключений
    conclusion = []
    for path in path_files: # открываем Заключения с анкетой для чтения данных
        if len(path_files) != 0:
            wbz = openpyxl.load_workbook(path, keep_vba=True)
            wsz = wbz.worksheets[0]     # анкета из заключения:
            # staff, department, full_name, last_name, birthday, birth_place, country, series_passport, 
            # number_passport, date_given, snils, inn, reg_address, live_address, phone, email, education
            form_view = [wsz['C4'].value, wsz['C5'].value, wsz['C6'].value, wsz['C7'].value, wsz['C8'].value,
                         'None','None', wsz['C9'].value, wsz['D9'].value, wsz['E9'].value, 'None', wsz['C10'].value,
                         'None', 'None', 'None', 'None', 'None']
            if len(wbz.sheetnames) > 1:
                wsz = wbz.worksheets[1]
                if str(wsz['K1'].value) == 'ФИО':     # если есть анкета, перезаписываем переменную form_view
                    form_view = [wsz['C3'].value, wsz['D3'].value, wsz['K3'].value, wsz['S3'].value, wsz['L3'].value,
                                 wsz['M3'].value, wsz['T3'].value, wsz['P3'].value, wsz['Q3'].value, wsz['R3'].value,
                                 wsz['U3'].value, wsz['V3'].value, wsz['N3'].value, wsz['O3'].value, wsz['Y3'].value,
                                 wsz['Z3'].value, wsz['X3'].value]
            wsz = wbz.worksheets[0] # Данные проверки: check_work_place, check_passport, check_debt, check_bankruptcy, 
            # check_bki, check_affiliation, check_internet, check_cronos, check_cross, resume, date_check, officer
            form_check = [f"{wsz['C11'].value} - {wsz['D11'].value}; {wsz['C12'].value} - "
                          f"{wsz['D12'].value}; {wsz['C13'].value} - {wsz['D13'].value}", wsz['C17'].value,
                          wsz['C18'].value, wsz['C19'].value, wsz['C20'].value, wsz['C21'].value, wsz['C22'].value,
                          f"{wsz['B14'].value}: {wsz['C14'].value}; {wsz['B15'].value}: {wsz['C15'].value}",
                          wsz['C16'].value, wsz['C23'].value, wsz['C24'].value, wsz['C25'].value]
            conclusion.append(tuple(check_types(form_view + form_check)))   # формируем данные для запроса
            wbz.close()   # Закрываем книгу Excel
    print('Получены данные из заключений')
    return conclusion

def create_link(sheet, subdir, row_num) -> None:    # создаем гиперссылки и перемещаем папки
    for n in row_num:
        for sub in subdir:
            if str(sheet['B' + str(n)].value.strip().lower()) == sub.strip().lower():
                sbd = sheet['B' + str(n)].value.strip()
                lnk = f"{DESTINATION + sbd[0][0]}\\{sbd} - {sheet['A' + str(n)].value}"
                sheet['L' + str(n)].hyperlink = lnk    # записывает в книгу
                shutil.move(WORK_DIR + sbd, lnk)
    print('Гиперссылки записаны, папки перемещены в архив')

def insert_db(query, value) -> None:  # запись в БД
    with sqlite3.connect(CONNECT, timeout=5.0) as con:
        cur = con.cursor()
        if len(value) > 0:
            cur.executemany(query, value)
    print('Данные переданы в БД')

def registry_check(row_num, sheet) -> None:  # Получаем данные из реестра и передаем в БД
    reg_val = []
    for n in row_num: # получаем список значений строк соответствующих сегодняшней дате
        reg_val.append(tuple(check_types([c.value for cell in sheet['B' + str(n):'L' + str(n)] for c in cell])))
    print('Получена информация из реестра')
    insert_db("INSERT INTO registry ({}) VALUES ({}?)".format(SQL_REG, '?, ' * (len(SQL_REG.split()) - 1)), reg_val)

def inquiry_check() -> None:  # Получаем данные из реестра запросов и передаем в БД
    wbc = openpyxl.load_workbook(INFO_FILE, keep_vba=True)
    wsc = wbc.worksheets[0]   # открываем первый лист книги MAIN_FILE для чтения и записи данных
    row_num = range_row(wsc['G1':'G2000'])
    if len(row_num) > 0:
        print('Данные за сегодня найдены')
        inquiry_val = []
        for n in row_num:  # получаем список значений строк соответствующих сегодняшней дате
            inquiry_val.append(tuple(check_types([c.value for cell in wsc['A' + str(n):'G' + str(n)] for c in cell])))
        print('Получена информация из реестра запросов')
        inquiry_query = "INSERT INTO inquiry ({}) VALUES ({}?)".format(SQL_INQ, '?, ' * (len(SQL_INQ.split()) - 1))
        insert_db(inquiry_query, inquiry_val)
    wbc.close()  # Закрываем книгу Excel

def cand_check() -> None:   # разбор реестра Кандидаты и Заключений
    wb = openpyxl.load_workbook(MAIN_FILE, keep_vba=True, read_only=False)
    ws = wb.worksheets[0]   # открываем первый лист книги MAIN_FILE для чтения и записи данных
    num_row = range_row(ws['K5000':'K20000']) # записываем номера строк соответствующих сегодняшней дате
    if len(num_row) == 0:
        print('Данные за сегодня отсутствуют')
        wb.close()  # Закрываем книгу Excel
    else:
        print('Получен список строк из книги')
        subdirectory = dir_range(ws, num_row)   # список директорий, которые соответствуют фамилиям кандидатов
        if len(subdirectory) > 0:
            print('Создан список папок из рабочей директории')
            path_files = file_range(subdirectory)
            if len(path_files) > 0:
                print('Файлы заключений найдены')
                ins_cand = "INSERT INTO candidates ({}) VALUES ({}?)".format(SQL_CAND, '?, ' * (len(SQL_CAND.split()) - 1))
                insert_db(ins_cand, parse_excel(path_files))  # передаем данные в БД
            create_link(ws, subdirectory, num_row)  # Создание гиперссылки и перемещение папок
        registry_check(num_row, ws)  # получаем данные из реестра и передаем в БД
        wb.save(MAIN_FILE)  # Сохраняем книгу Excel

def main(): # проверка файлов на изменения с датой сегодня, тип запуска программы в зависимости от результата
    main_file_date = date.fromtimestamp(os.path.getmtime(MAIN_FILE))
    info_file_date = date.fromtimestamp(os.path.getmtime(INFO_FILE))
    if main_file_date == DATE or info_file_date == DATE:
        shutil.copy(CONNECT, DESTINATION)  # backup files
    if main_file_date == DATE:
        shutil.copy(MAIN_FILE, DESTINATION)
        cand_check()
    if info_file_date == DATE:
        shutil.copy(INFO_FILE, DESTINATION)
        inquiry_check()

if __name__ == "__main__":
    main()
