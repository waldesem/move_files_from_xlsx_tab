import openpyxl
import os.path
import shutil
from datetime import date

#сегодняшняя дата
date_today = date.today().strftime('%Y-%m-%d') + ' 00:00:00'

#главный файл кандидатов
main_file = r'\\cronosx1\New folder\УВБ\Отдел корпоративной защиты\Кандидаты\Кандидаты.xlsm'
#рабочая папка кандидатов
output = r'\\cronosx1\New folder\УВБ\Отдел корпоративной защиты\Кандидаты\\'
#место хранения отработанных кандидатов
dest = r'\\cronosx1\New folder\УВБ\Отдел корпоративной защиты\Персонал\Персонал-2\\'

print('Начинаем работу... Создаем резервную копию книги Exel в', dest)
shutil.copy(main_file, dest)

print('Открываем книгу по адресу:', main_file, 'для чтения и записи данных')
wb = openpyxl.load_workbook(main_file, keep_vba=True)
#Берем первый лист
ws = wb.worksheets[0]
print('Идет поиск ячеек с датами согласования сегодня') #(огранbчение 30 тыс. строк)
col_range = ws['K1':'K30000']
for cell in col_range:
    for c in cell:
        #если запись в ячейке равна сегодняшней дате
        if str(c.value) == date_today:
            print('Найдены записи с сегодняшней датой', date_today)
            #берем номер строки
            row_num = c.row
            #берем значение из ячейки фамилия, которое советствууюет номеру строки дата
            fio = ws['B'+str(row_num)].value
            print('Получена запись с ФИО кандидата:', fio)
            #Перебираем каталоги в исходной папке
            for dirs, subdirs, files in os.walk(output):
                for subdir in subdirs:
                    #если имя папки такое же как и значение в ячейке фамилия
                    if subdir.lower().rstrip() == fio.lower().rstrip():
                        print('Найдена папка,', subdir, 'которая соответствует ФИО кандидата')
                        #разбираем посимвольно имя папки
                        letter = [i for i in fio]
                        #создаем ссылку для целевой деректории с первым именем буквы по алфавиту, добавляем к имени уникальный ID
                        hlink = dest+'\\'+letter[0]+'\\'+subdir+' - '+str(ws['A'+str(row_num)].value)
                        print('Создана гиперссылка:', hlink)
                        #переносим папку из исходной в целевую папку
                        shutil.move(output+subdir, hlink)
                        #добавляем в файл книги гиперссылку, куда помещена папка
                        ws['L'+str(row_num)].hyperlink = hlink
                    else:
                        continue
        else:
            continue                
print('Cохраняем книгу Excel')
wb.save(main_file)
